from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from app import db
from app.models.models import VerificationTransaction, Customer, AuditLog, User
from app.services.audit_service import log_action
from datetime import datetime, timedelta
import os, io

reports_bp = Blueprint('reports', __name__)

def require_role(*roles):
    from functools import wraps
    from flask_jwt_extended import get_jwt_identity
    def decorator(f):
        @wraps(f)
        @jwt_required()
        def wrapper(*args, **kwargs):
            user = User.query.get(get_jwt_identity())
            if not user or user.role not in roles:
                return jsonify({'error': 'Access denied'}), 403
            return f(*args, **kwargs)
        return wrapper
    return decorator

@reports_bp.route('/daily', methods=['GET'])
@jwt_required()
def daily_report_data():
    """FR-020: Daily fraud detection summary"""
    date_str = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        target_date = datetime.utcnow().date()
    
    txns = VerificationTransaction.query.filter(
        db.func.date(VerificationTransaction.created_at) == target_date
    ).all()
    
    total = len(txns)
    genuine = sum(1 for t in txns if t.classification == 'Genuine')
    suspected = sum(1 for t in txns if t.classification == 'Suspected Forgery')
    highly_sus = sum(1 for t in txns if t.classification == 'Highly Suspicious')
    
    avg_score = sum(t.match_score for t in txns) / total if total else 0
    avg_time = sum(t.processing_time_ms or 0 for t in txns) / total if total else 0
    
    # Branch breakdown
    branches = {}
    for t in txns:
        b = t.branch or 'Unknown'
        if b not in branches:
            branches[b] = {'total': 0, 'genuine': 0, 'suspected': 0, 'highly_suspicious': 0}
        branches[b]['total'] += 1
        if t.classification == 'Genuine':
            branches[b]['genuine'] += 1
        elif t.classification == 'Suspected Forgery':
            branches[b]['suspected'] += 1
        else:
            branches[b]['highly_suspicious'] += 1
    
    return jsonify({
        'date': date_str,
        'total': total,
        'genuine': genuine,
        'suspected_forgery': suspected,
        'highly_suspicious': highly_sus,
        'avg_match_score': round(avg_score, 2),
        'avg_processing_time_ms': round(avg_time, 1),
        'branches': branches,
        'transactions': [t.to_dict() for t in txns[:100]]
    })

@reports_bp.route('/monthly', methods=['GET'])
@jwt_required()
def monthly_report_data():
    """FR-020: Monthly fraud detection summary"""
    year = int(request.args.get('year', datetime.utcnow().year))
    month = int(request.args.get('month', datetime.utcnow().month))
    
    start = datetime(year, month, 1)
    if month == 12:
        end = datetime(year + 1, 1, 1)
    else:
        end = datetime(year, month + 1, 1)
    
    txns = VerificationTransaction.query.filter(
        VerificationTransaction.created_at >= start,
        VerificationTransaction.created_at < end
    ).all()
    
    total = len(txns)
    genuine = sum(1 for t in txns if t.classification == 'Genuine')
    suspected = sum(1 for t in txns if t.classification == 'Suspected Forgery')
    highly_sus = sum(1 for t in txns if t.classification == 'Highly Suspicious')
    
    # Daily breakdown
    daily = {}
    for t in txns:
        d = t.created_at.strftime('%Y-%m-%d')
        if d not in daily:
            daily[d] = {'total': 0, 'genuine': 0, 'suspected': 0, 'highly_suspicious': 0}
        daily[d]['total'] += 1
        if t.classification == 'Genuine':
            daily[d]['genuine'] += 1
        elif t.classification == 'Suspected Forgery':
            daily[d]['suspected'] += 1
        else:
            daily[d]['highly_suspicious'] += 1
    
    return jsonify({
        'year': year,
        'month': month,
        'total': total,
        'genuine': genuine,
        'suspected_forgery': suspected,
        'highly_suspicious': highly_sus,
        'daily_breakdown': daily
    })

@reports_bp.route('/audit-trail', methods=['GET'])
@require_role('admin', 'auditor')
def audit_trail():
    """FR-023: Auditor read-only audit trail access"""
    user_id = get_jwt_identity()
    
    limit = int(request.args.get('limit', 100))
    action_filter = request.args.get('action', '')
    
    query = AuditLog.query.order_by(AuditLog.created_at.desc())
    if action_filter:
        query = query.filter(AuditLog.action.ilike(f'%{action_filter}%'))
    
    logs = query.limit(limit).all()
    log_action(user_id, 'VIEW_AUDIT_TRAIL', 'audit_log')
    return jsonify([l.to_dict() for l in logs])

@reports_bp.route('/export/pdf', methods=['GET'])
@jwt_required()
def export_pdf():
    """FR-021: PDF report export"""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    
    user_id = get_jwt_identity()
    date_str = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        target_date = datetime.utcnow().date()
    
    txns = VerificationTransaction.query.filter(
        db.func.date(VerificationTransaction.created_at) == target_date
    ).all()
    
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    elements.append(Paragraph('FSDS Daily Report', styles['Title']))
    elements.append(Paragraph(f'WeCare Software Solutions LTD. | Date: {date_str}', styles['Normal']))
    elements.append(Spacer(1, 0.5*cm))
    
    # Summary table
    total = len(txns)
    genuine = sum(1 for t in txns if t.classification == 'Genuine')
    suspected = sum(1 for t in txns if t.classification == 'Suspected Forgery')
    highly_sus = sum(1 for t in txns if t.classification == 'Highly Suspicious')
    
    summary_data = [
        ['Metric', 'Count'],
        ['Total Verifications', str(total)],
        ['Genuine', str(genuine)],
        ['Suspected Forgery', str(suspected)],
        ['Highly Suspicious', str(highly_sus)],
    ]
    summary_table = Table(summary_data, colWidths=[8*cm, 4*cm])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.5*cm))
    
    # Transactions table
    if txns:
        elements.append(Paragraph('Transaction Details', styles['Heading2']))
        txn_data = [['Transaction ID', 'Customer', 'Score', 'Classification', 'Officer', 'Time']]
        for t in txns[:50]:
            txn_data.append([
                t.transaction_id,
                t.customer.full_name if t.customer else 'N/A',
                f'{t.match_score:.1f}%',
                t.classification,
                t.officer.full_name if t.officer else 'N/A',
                t.created_at.strftime('%H:%M:%S')
            ])
        txn_table = Table(txn_data, colWidths=[3.5*cm, 3*cm, 1.8*cm, 3.5*cm, 3*cm, 2*cm])
        txn_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5c')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTSIZE', (0, 0), (-1, -1), 7),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f4f8')]),
        ]))
        elements.append(txn_table)
    
    doc.build(elements)
    buf.seek(0)
    
    log_action(user_id, 'EXPORT_PDF_REPORT', 'report', None, f'Date: {date_str}')
    return send_file(buf, mimetype='application/pdf', as_attachment=True,
                     download_name=f'FSDS_Report_{date_str}.pdf')

@reports_bp.route('/export/excel', methods=['GET'])
@jwt_required()
def export_excel():
    """FR-021: Excel report export"""
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    
    user_id = get_jwt_identity()
    date_str = request.args.get('date', datetime.utcnow().strftime('%Y-%m-%d'))
    
    try:
        target_date = datetime.strptime(date_str, '%Y-%m-%d').date()
    except:
        target_date = datetime.utcnow().date()
    
    txns = VerificationTransaction.query.filter(
        db.func.date(VerificationTransaction.created_at) == target_date
    ).all()
    
    wb = Workbook()
    ws = wb.active
    ws.title = 'Daily Report'
    
    header_fill = PatternFill(start_color='1a3a5c', end_color='1a3a5c', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    
    headers = ['Transaction ID', 'Customer Name', 'Account Number', 
               'Match Score (%)', 'Classification', 'Document Type',
               'Officer', 'Branch', 'Processing Time (ms)', 'Timestamp']
    
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    for row, t in enumerate(txns, 2):
        ws.cell(row=row, column=1, value=t.transaction_id)
        ws.cell(row=row, column=2, value=t.customer.full_name if t.customer else '')
        ws.cell(row=row, column=3, value=t.customer.account_number if t.customer else '')
        ws.cell(row=row, column=4, value=round(t.match_score, 2))
        ws.cell(row=row, column=5, value=t.classification)
        ws.cell(row=row, column=6, value=t.document_type or '')
        ws.cell(row=row, column=7, value=t.officer.full_name if t.officer else '')
        ws.cell(row=row, column=8, value=t.branch)
        ws.cell(row=row, column=9, value=t.processing_time_ms)
        ws.cell(row=row, column=10, value=t.created_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    
    log_action(user_id, 'EXPORT_EXCEL_REPORT', 'report', None, f'Date: {date_str}')
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f'FSDS_Report_{date_str}.xlsx')
